#!/usr/bin/env python3
"""
gsheet_sync.py — синхронизация дашборда с Google-таблицей управленческой отчётности.

Источник — книга, выгруженная из Google Sheets как xlsx (по одному ЛИСТУ на дату,
имя листа = ДД.ММ.ГГГГ, колонки: Адрес, ПВЗ, Дата, Кол-во заказов, Общая сумма).
Это ровно тот формат, что ест build_data.py, поэтому накопление и вшивание в дашборд
выполняет именно он (полные недели не трогает, последнюю неполную дополняет,
новые дописывает; правит ТОЛЬКО листы «Заказы_по_неделям» / «Объём_по_неделям»).

Этот скрипт добавляет сверху:
  • детект НОВЫХ дат (сверка с леджером .gsheet_ledger.json),
  • защиту от залоченного эталона (открыт в Excel),
  • обновление леджера только при успешном прогоне.

Запуск:  python gsheet_sync.py [__gsheet.xlsx] [Эталон_ПВЗ.xlsx] [index.html]
Предусловие: Эталон_ПВЗ.xlsx должен быть ЗАКРЫТ в Excel.
"""
import sys, os, json, subprocess, datetime as dt
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC  = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, '__gsheet.xlsx')
REF  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(HERE, 'Эталон_ПВЗ.xlsx')
HTML = sys.argv[3] if len(sys.argv) > 3 else os.path.join(HERE, 'index.html')
LEDGER = os.path.join(HERE, '.gsheet_ledger.json')


def parse_date(s):
    try:
        return dt.datetime.strptime(str(s).strip(), '%d.%m.%Y').date()
    except ValueError:
        return None


def main():
    if not os.path.exists(SRC):
        raise SystemExit(f'Нет файла выгрузки: {SRC}. Сначала скачай Google-таблицу как xlsx.')

    # 1. даты выгрузки = имена листов
    wb = load_workbook(SRC, read_only=True)
    sheet_dates = {n.strip(): parse_date(n) for n in wb.sheetnames}
    sheet_dates = {n: d for n, d in sheet_dates.items() if d}
    wb.close()
    if not sheet_dates:
        raise SystemExit('В выгрузке нет листов с датой ДД.ММ.ГГГГ — нечего обновлять.')
    all_dates = sorted(sheet_dates, key=lambda s: sheet_dates[s])

    # 2. леджер -> новые даты
    first_run = not os.path.exists(LEDGER)
    prev = set()
    if not first_run:
        try:
            prev = set(json.load(open(LEDGER, encoding='utf-8')))
        except (ValueError, OSError):
            first_run = True
    new_dates = [s for s in all_dates if s not in prev]
    if first_run:
        print(f'ЛЕДЖЕР: первичная инициализация, в выгрузке {len(all_dates)} дат '
              f'({all_dates[0]} … {all_dates[-1]}). Впредь буду показывать только новые.')
    elif new_dates:
        print(f'НОВЫЕ ДАТЫ ({len(new_dates)}): ' + ', '.join(new_dates))
    else:
        print('НОВЫЕ ДАТЫ: нет — всё уже учтено. (Последняя неполная неделя всё равно будет пересчитана.)')

    # 3. эталон должен быть доступен на запись (иначе build_data упадёт на сохранении)
    try:
        with open(REF, 'r+b'):
            pass
    except FileNotFoundError:
        raise SystemExit(f'Нет эталона: {REF}')
    except PermissionError:
        raise SystemExit(f'{REF} занят (открыт в Excel?). Закрой файл и повтори — данные не тронуты.')

    # 4. проверенное накопление + вшивание в дашборд
    print('--- build_data.py ---')
    r = subprocess.run([sys.executable, os.path.join(HERE, 'build_data.py'), SRC, REF, HTML])
    if r.returncode != 0:
        raise SystemExit(f'build_data.py завершился с ошибкой (код {r.returncode}). Леджер НЕ обновлён.')

    # 5. обновить леджер только при успехе
    json.dump(all_dates, open(LEDGER, 'w', encoding='utf-8'), ensure_ascii=False)
    print(f'Леджер обновлён: учтено дат = {len(all_dates)} ({all_dates[0]} … {all_dates[-1]}).')


if __name__ == '__main__':
    main()
