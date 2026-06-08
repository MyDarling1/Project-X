#!/usr/bin/env python3
"""
build_data.py — обновляет эталонный Excel и данные дашборда из дневного узум.xlsx.

ПРИНЦИП НАКОПЛЕНИЯ (важно):
  • Эталон НАКАПЛИВАЕТ недели. Старые ПОЛНЫЕ недели (7 дней) НЕ переписываются.
  • Только ПОСЛЕДНЯЯ неполная неделя обновляется при каждом прогоне (в неё дотекают дни).
    Как только она становится полной — фиксируется.
  • Новые недели дописываются справа.
  • Лист «Справочник» (сегмент/комиссия/опекс/УК/гарантия/прогнозы) — только читается.
  • Дневной узум.xlsx — только читается.

Дашборд: страница «Недельная история» показывает ВСЕ накопленные недели (с прокруткой);
«Обзор» (run-rate, безубыток) использует последнюю ПОЛНУЮ неделю.

Запуск:  python build_data.py [узум.xlsx] [Эталон_ПВЗ.xlsx] [index.html]
Затем:   git add -A && git commit -m "data update" && git push   (Railway пересоберёт)
"""
import sys, os, json, re, datetime as dt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

DAILY = sys.argv[1] if len(sys.argv)>1 else 'узум.xlsx'
REF   = sys.argv[2] if len(sys.argv)>2 else 'Эталон_ПВЗ.xlsx'
HTML  = sys.argv[3] if len(sys.argv)>3 else 'index.html'

NAVY='1F3864'; BLUE='2F5496'; WHITE='FFFFFF'
thin=Side(style='thin',color='BFBFBF'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
def monday(d): return d - dt.timedelta(days=d.weekday())
def wlabel(m): e=m+dt.timedelta(days=6); return f'{m.day:02d}.{m.month:02d}–{e.day:02d}.{e.month:02d}'

def aggregate(path, ref=None):
    """Дневной файл -> {monday: {code: [ord_sum, rev_sum, days]}}, всё что есть в файле.
    days = ДНИ РАБОТЫ ПВЗ (день >= даты старта из Справочника): в выгрузке точка присутствует
    и до открытия (со строкой 0 заказов), такие дни в знаменатель не идут."""
    x=pd.read_excel(path, sheet_name=None)
    req={'Адрес','ПВЗ','Дата','Кол-во заказов','Общая сумма'}
    starts={}
    if ref:
        for c,rr in ref.items():
            try: starts[c]=dt.date.fromisoformat(str(rr.get('start'))[:10])
            except (ValueError, TypeError): pass
    acc={}
    for name,df in x.items():
        try: day=dt.datetime.strptime(name.strip(),'%d.%m.%Y').date()
        except ValueError: continue
        df=df.copy(); df.columns=[str(c).strip() for c in df.columns]
        if not req.issubset(set(df.columns)):
            raise SystemExit(f'Лист "{name}": не хватает колонок. Есть: {list(df.columns)}')
        m=monday(day)
        for _,r in df.iterrows():
            code=str(r['ПВЗ']).strip()
            if not code or code=='nan': continue
            o=r['Кол-во заказов']
            if pd.isna(o): continue
            try: o=float(o)
            except (ValueError, TypeError): continue   # нечисловая ячейка заказов ('-', текст, мусор) — пропускаем, не валим всю сборку
            s=r['Общая сумма']
            try: s=float(s) if pd.notna(s) else 0.0
            except (ValueError, TypeError): s=0.0
            d=acc.setdefault(m,{}).setdefault(code,[0.0,0.0,set()])
            d[0]+=o; d[1]+=s
            st=starts.get(code)
            if st is None or day>=st: d[2].add(day)   # день идёт в знаменатель только с даты старта ПВЗ
    if not acc: raise SystemExit('Не найдено листов с датой ДД.ММ.ГГГГ')
    return acc

def read_ref(path):
    wb=load_workbook(path, data_only=True); sh=wb['Справочник']
    cols={str(sh.cell(4,j).value).strip():j for j in range(1,sh.max_column+1)}
    need=['Код','AKA','Город','Сегмент','Комиссия','Аренда, сум/мес','ЗП, сум/мес','Затраты на УК, $/мес',
          'Разное, сум/мес','Гарантия, сум/мес','Прогноз 1-го мес, зак/день','Яндекс-пик, зак/день','Старт']
    for n in need:
        if n not in cols: raise SystemExit(f'Справочник: нет колонки «{n}»')
    out={}
    for rr in range(5, sh.max_row+1):
        code=sh.cell(rr,cols['Код']).value
        if not code: continue
        g=lambda k: sh.cell(rr,cols[k]).value
        st=g('Старт'); st=st.date() if hasattr(st,'date') else st
        rent=int(round(float(g('Аренда, сум/мес')))); sal=int(round(float(g('ЗП, сум/мес'))))
        misc=int(round(float(g('Разное, сум/мес') or 0)))
        out[str(code).strip()]=dict(aka=g('AKA'),city=g('Город'),seg=int(g('Сегмент')),
            comm=float(g('Комиссия')),rentSum=rent,salarySum=sal,miscSum=misc,
            uk=round(float(g('Затраты на УК, $/мес')),2),guarFull=int(g('Гарантия, сум/мес')),
            fcMonth1=int(g('Прогноз 1-го мес, зак/день')),yandex=int(g('Яндекс-пик, зак/день')),start=str(st))
    return out

def read_archive(wb, sheet):
    """Лист-архив -> (weeks[monday], vals{monday:{code:val}}, dc{monday:int}). Пустой -> ({},{},{})."""
    if sheet not in wb.sheetnames: return [],{},{}
    sh=wb[sheet]
    keys={}
    for j in range(3, sh.max_column+1):
        v=sh.cell(3,j).value
        if v: keys[j]=dt.date.fromisoformat(str(v))
    if not keys: return [],{},{}
    dc={}; 
    for j,m in keys.items():
        d=sh.cell(5,j).value; dc[m]=int(d) if d is not None else 0
    # найти строку «В среднем на ПВЗ»
    avg_r=None
    for rr in range(6, sh.max_row+1):
        if str(sh.cell(rr,1).value).strip()=='В среднем на ПВЗ': avg_r=rr; break
    end=(avg_r-1) if avg_r else sh.max_row
    vals={m:{} for m in keys.values()}
    for rr in range(6, end+1):
        code=sh.cell(rr,1).value
        if not code: continue
        for j,m in keys.items():
            v=sh.cell(rr,j).value
            if isinstance(v,(int,float)): vals[m][str(code).strip()]=v
    weeks=sorted(keys.values())
    return weeks, vals, dc

def merge(arch_w, arch_v, arch_dc, new_acc, pick):
    """Накопление. pick(code,monday)->значение из new_acc. Возвращает (weeks, vals, dc)."""
    new_w=sorted(new_acc.keys())
    old_trailing=arch_w[-1] if arch_w else None
    new_dc={m: len(next(iter(new_acc[m].values()))[2]) if new_acc[m] else 0 for m in new_w}
    # точнее: дни недели = объединение дней всех точек
    for m in new_w:
        days=set()
        for code,(o,s,dd) in new_acc[m].items(): days|=dd
        new_dc[m]=len(days)
    weeks=sorted(set(arch_w)|set(new_w))
    vals={}; dc={}
    for m in weeks:
        in_arch = m in arch_dc
        # последняя неполная неделя дотекает днями, но НЕ должна сжиматься: берём новое только если дней не меньше, чем уже накоплено
        trailing_update = (m==old_trailing and in_arch and arch_dc[m]<7 and m in new_acc and new_dc.get(m,0)>=arch_dc[m])
        if trailing_update:
            vals[m]={code: pick(code,m) for code in new_acc[m]}; dc[m]=new_dc[m]
        elif in_arch:
            # зафиксированная полная неделя ИЛИ хвост, который иначе бы сжался (новый файл короче) — сохраняем архив
            if m==old_trailing and arch_dc[m]<7 and m in new_acc and new_dc.get(m,0)<arch_dc[m]:
                print(f'  [накопление] неделя {wlabel(m)}: новый файл короче ({new_dc.get(m,0)} дн.) архива ({arch_dc[m]} дн.) — оставляю накопленное', file=sys.stderr)
            vals[m]=dict(arch_v.get(m,{})); dc[m]=arch_dc[m]
        elif m in new_acc:
            # совершенно новая неделя
            vals[m]={code: pick(code,m) for code in new_acc[m]}; dc[m]=new_dc[m]
        else:
            vals[m]=dict(arch_v.get(m,{})); dc[m]=arch_dc.get(m,0)
    return weeks, vals, dc

def write_archive(wb, sheet, title, weeks, vals, dc, codes, ref, fmt):
    if sheet in wb.sheetnames: 
        sh=wb[sheet]; sh.delete_rows(1, sh.max_row+1)
    else: sh=wb.create_sheet(sheet)
    sh['A1']=title; sh['A1'].font=Font(name='Arial',bold=True,size=12,color=NAVY)
    sh['A2']='Накопительный архив: полные недели не меняются, последняя неполная — обновляется. Звёздочка = неполная неделя.'
    sh['A2'].font=Font(name='Arial',italic=True,size=9,color='595959')
    # row3: ключи (ISO), скрытая
    for i,m in enumerate(weeks): sh.cell(3,3+i,m.isoformat())
    sh.row_dimensions[3].hidden=True
    # row4: заголовки
    hdr=['Код','AKA']+[wlabel(m)+('*' if dc[m]<7 else '') for m in weeks]
    for j,h in enumerate(hdr,1):
        c=sh.cell(4,j,h); c.font=Font(name='Arial',bold=True,color=WHITE,size=10)
        c.fill=PatternFill('solid',fgColor=NAVY); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BD
    # row5: дней с данными
    c=sh.cell(5,1,'Дней с данными'); c.font=Font(name='Arial',size=9,italic=True,color='595959'); c.border=BD
    sh.cell(5,2,'').border=BD
    for i,m in enumerate(weeks):
        c=sh.cell(5,3+i,dc[m]); c.number_format='0'; c.font=Font(name='Arial',size=9,italic=True,color='595959'); c.alignment=Alignment(horizontal='center'); c.border=BD
    # данные
    rr=6
    for code in codes:
        sh.cell(rr,1,code).border=BD; sh.cell(rr,1).font=Font(name='Arial',size=10)
        sh.cell(rr,2,ref.get(code,{}).get('aka','')).border=BD; sh.cell(rr,2).font=Font(name='Arial',size=10)
        for i,m in enumerate(weeks):
            v=vals[m].get(code, 0)
            c=sh.cell(rr,3+i, round(v,2) if fmt=='0.0' else int(round(v)))
            c.number_format=fmt; c.font=Font(name='Arial',size=10); c.alignment=Alignment(horizontal='center'); c.border=BD
        rr+=1
    last=rr-1
    c=sh.cell(rr,1,'В среднем на ПВЗ'); c.font=Font(name='Arial',bold=True,color=WHITE,size=10); c.fill=PatternFill('solid',fgColor=BLUE); c.border=BD
    c2=sh.cell(rr,2,''); c2.fill=PatternFill('solid',fgColor=BLUE); c2.border=BD
    for i in range(len(weeks)):
        L=get_column_letter(3+i)
        c=sh.cell(rr,3+i,f'=AVERAGE({L}6:{L}{last})'); c.number_format=fmt
        c.font=Font(name='Arial',bold=True,color=WHITE,size=10); c.fill=PatternFill('solid',fgColor=BLUE); c.alignment=Alignment(horizontal='center'); c.border=BD
    for i,(col,w) in enumerate(zip(['A','B'],[11,20])): sh.column_dimensions[col].width=w
    for i in range(len(weeks)): sh.column_dimensions[get_column_letter(3+i)].width=13
    sh.freeze_panes='C6'

def read_payouts(path, year):
    """Лист «выплаты » (помесячный ФАКТ выплат зарплат по ПВЗ) -> (isoKey 'YYYY-MM', {code: сум}) или None.
    Шапка может быть не в 1-й строке; над ней — метка периода вида 'DD-DD.MM'. Джойн по коду ПВЗ."""
    wb=load_workbook(path, read_only=True, data_only=True)
    target=None
    for n in wb.sheetnames:
        if str(n).strip().lower()=='выплаты': target=n; break
    if not target: wb.close(); return None
    rows=[list(r) for r in wb[target].iter_rows(values_only=True)]
    wb.close()
    hdr_i=None
    for i,r in enumerate(rows):
        cells=[str(c).strip() if c is not None else '' for c in r]
        if 'ПВЗ' in cells and any(str(c).strip().startswith('Общая сумма') for c in cells): hdr_i=i; break
    if hdr_i is None: return None
    hdr=[str(c).strip() if c is not None else '' for c in rows[hdr_i]]
    pvz_j=hdr.index('ПВЗ')
    sum_j=next(j for j,c in enumerate(hdr) if c.startswith('Общая сумма'))
    month=None
    for i in range(hdr_i,-1,-1):
        for c in rows[i]:
            if c is None: continue
            mm=re.search(r'(\d{1,2})\.(\d{1,2})\s*$', str(c).strip())
            if mm: month=int(mm.group(2)); break
        if month: break
    if not month: return None
    key=f'{year:04d}-{month:02d}'; sal={}
    for r in rows[hdr_i+1:]:
        if pvz_j>=len(r) or sum_j>=len(r): continue
        code=r[pvz_j]; amt=r[sum_j]
        if code is None or str(code).strip()=='' or str(code).strip().lower()=='nan': continue
        try: amt=float(amt)
        except (TypeError,ValueError): continue
        sal[str(code).strip()]=round(amt)
    return (key, sal) if sal else None

def main():
    ref=read_ref(REF)
    acc=aggregate(DAILY, ref)
    wb=load_workbook(REF)
    # архивы
    aw,av,adc=read_archive(wb,'Заказы_по_неделям')
    weeks,ord_vals,dc = merge(aw,av,adc,acc, lambda c,m: acc[m][c][0]/max(len(acc[m][c][2]),1))  # заказы/день
    _,_,_ = (aw,av,adc)
    bw,bv,bdc=read_archive(wb,'Объём_по_неделям')
    _,vol_vals,_ = merge(bw,bv,bdc,acc, lambda c,m: acc[m][c][1])  # суммы объёма
    cw,cv,cdc=read_archive(wb,'Дни_по_неделям')
    weeks_d,day_vals,dc_d = merge(cw,cv,cdc,acc, lambda c,m: len(acc[m][c][2]))  # #5: дни с данными по каждому ПВЗ за неделю (накопительно)
    def pdays(c,m): return day_vals[m].get(c,0) if m in day_vals else dc[m]       # дни этого ПВЗ за неделю; для недель без архива дней — союз дней недели (фолбэк)
    codes=[c for c in ref]  # порядок как в справочнике
    # отсортируем по последней неделе (заказы/день) убыв.
    last_m=weeks[-1]
    codes.sort(key=lambda c:-ord_vals[last_m].get(c,0))
    write_archive(wb,'Заказы_по_неделям','Средние заказы в день по неделям (накопительно)',weeks,ord_vals,dc,codes,ref,'0.0')
    write_archive(wb,'Объём_по_неделям','Сумма выручки (объём) по неделям, сум (накопительно)',weeks,vol_vals,dc,codes,ref,'#,##0')
    write_archive(wb,'Дни_по_неделям','Дней с данными по каждому ПВЗ за неделю (накопительно)',weeks_d,day_vals,dc_d,codes,ref,'0')
    tmp_ref = REF + '.tmp'; wb.save(tmp_ref); os.replace(tmp_ref, REF)   # атомарная запись (без частичной порчи)
    print(f'{REF}: накоплено недель = {len(weeks)} ({wlabel(weeks[0])} … {wlabel(weeks[-1])}); дней посл. недели = {dc[weeks[-1]]}')

    # ---- данные дашборда из ПОСЛЕДНИХ 4 недель ----
    last4=weeks[-4:]
    full_runIdx=max([i for i,m in enumerate(weeks) if dc[m]>=7], default=len(weeks)-1)
    runM=weeks[full_runIdx]
    at=lambda lst,i: lst[i] if i<len(lst) else 0
    DATA=[]; WKpvz={}
    for code in codes:
        if code not in ref: continue
        ordd=[round(ord_vals[m].get(code,0),2) for m in last4]          # последние 4 — для карточки точки
        ordFull=[round(ord_vals[m].get(code,0),2) for m in weeks]        # ВСЯ история — для страницы истории
        revFull=[round(vol_vals[m].get(code,0)/max(pdays(code,m),1)) for m in weeks]   # #5: оборот/день делим на дни именно этого ПВЗ
        runRate=round(ord_vals[runM].get(code,0),2)
        revRun=round(vol_vals[runM].get(code,0)/max(pdays(code,runM),1))
        tot_o=sum(ord_vals[m].get(code,0)*pdays(code,m) for m in weeks)                # истинные суммарные заказы (заказы/день × дни ПВЗ)
        tot_r=sum(vol_vals[m].get(code,0) for m in weeks)
        tot_d=sum(pdays(code,m) for m in weeks)                                        # суммарные активные дни именно этого ПВЗ
        r=ref[code]
        DATA.append(dict(code=code,aka=r['aka'],city=r['city'],seg=r['seg'],comm=r['comm'],
            rentSum=r['rentSum'],salarySum=r['salarySum'],miscSum=r['miscSum'],uk=r['uk'],
            opex_usd=round((r['rentSum']+r['salarySum'])/12000,2),
            w1=at(ordd,0),w2=at(ordd,1),w3=at(ordd,2),w4=at(ordd,3),
            runRate=runRate,periodAvg=round(tot_o/tot_d,2) if tot_d else 0,
            fcMonth1=r['fcMonth1'],yandex=r['yandex'],revW3=revRun,
            revPeriod=round(tot_r/tot_d) if tot_d else 0,avgCheck=round(tot_r/tot_o) if tot_o else 0,
            gmvFc_usd=round(r['fcMonth1']*14),guarFull=r['guarFull'],start=r['start']))
        WKpvz[code]=dict(aka=r['aka'],city=r['city'],seg=r['seg'],ord=ordFull,rev=revFull,days=[pdays(code,m) for m in weeks])
    DATA.sort(key=lambda d:-d['revW3'])
    net_ord=[round(sum(WKpvz[c]['ord'][i] for c in WKpvz),1) for i in range(len(weeks))]
    net_rev=[sum(WKpvz[c]['rev'][i] for c in WKpvz) for i in range(len(weeks))]
    WK=dict(weeks=[wlabel(m) for m in weeks],years=[m.year for m in weeks],runIdx=full_runIdx,dayCounts=[dc[m] for m in weeks],
            net=dict(ord=net_ord,rev=net_rev),pvz=WKpvz)
    html=open(HTML,encoding='utf-8').read()
    html=re.sub(r'const DATA = \[.*?\];', lambda _: 'const DATA = '+json.dumps(DATA,ensure_ascii=False)+';', html,count=1,flags=re.S)
    html=re.sub(r'const WK = \{.*?\};\n', lambda _: 'const WK = '+json.dumps(WK,ensure_ascii=False)+';\n', html,count=1,flags=re.S)
    # ---- факт выплат зарплат (лист «выплаты ») → накопительный const SAL в дашборде ----
    pay=read_payouts(DAILY, weeks[-1].year)
    mex=re.search(r'const SAL = (\{.*?\});', html, flags=re.S)
    try: SAL=json.loads(mex.group(1)) if mex else {}
    except ValueError: SAL={}
    if pay:
        key,sal=pay; SAL[key]=sal
        print(f'{HTML}: факт зарплат {key} — ПВЗ={len(sal)}, итог={sum(sal.values()):,} сум')
    else:
        print(f'{HTML}: лист «выплаты » не найден в источнике — const SAL без изменений ({len(SAL)} мес.)')
    js_sal='const SAL = '+json.dumps(SAL,ensure_ascii=False)+';'
    if mex: html=re.sub(r'const SAL = \{.*?\};', lambda _: js_sal, html, count=1, flags=re.S)
    else:   html=re.sub(r'(const WK = \{.*?\};\n)', lambda mo: mo.group(1)+js_sal+'\n', html, count=1, flags=re.S)
    open(HTML,'w',encoding='utf-8').write(html)
    print(f'{HTML}: история — все {len(weeks)} нед.; Обзор run-rate=неделя {wlabel(runM)}, дни {WK["dayCounts"]}')
    print('Готово. Запусти recalc для эталона при желании, затем git commit & push.')

if __name__=='__main__': main()
