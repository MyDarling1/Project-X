#!/usr/bin/env python3
"""
build_satori.py — вносит ФАКТ из месячного отчёта «Сатори» (агентское вознаграждение Uzum)
в эталон и дашборд.

ПРИНЦИП (согласовано):
  • Источник: docx «Отчёт о выполнении агентского поручения» (UZUM ↔ SATORI IT).
  • Доход ПВЗ (факт) за месяц = СУММА всех строк ПВЗ: фиксированное вознаграждение
    (гарантия, прорейчено по дням) + п.1.8 + компенсация п.5.2.1.
  • Накопление по месяцам: лист «Сатори_факт» в эталоне; закрытые месяцы не переписываются,
    повторный прогон того же месяца — обновляет его.
  • Затраты/факт-P&L считаются в дашборде: доход − (аренда+ЗП+УК+разное)×активные_дни/дней_в_месяце.
    Активные дни берутся из периода строки отчёта.
  • Деньги — только из Сатори; заказы/оборот — из дневного файла; статика — из Справочника.

Запуск:  python build_satori.py "сатори_ГГГГ-ММ.docx" Эталон_ПВЗ.xlsx index.html
Затем:   git add -A && git commit -m "satori: <месяц>" && git push
"""
import sys, os, re, json, calendar, datetime as dt
import zipfile, xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

DOCX = sys.argv[1] if len(sys.argv) > 1 else 'сатори.docx'
REF  = sys.argv[2] if len(sys.argv) > 2 else 'Эталон_ПВЗ.xlsx'
HTML = sys.argv[3] if len(sys.argv) > 3 else 'index.html'

W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
NAVY = '1F3864'; BLUE = '2F5496'; WHITE = 'FFFFFF'
thin = Side(style='thin', color='BFBFBF'); BD = Border(left=thin, right=thin, top=thin, bottom=thin)
MONTHS = {'январ': 1, 'феврал': 2, 'март': 3, 'апрел': 4, 'ма': 5, 'июн': 6, 'июл': 7,
          'август': 8, 'сентябр': 9, 'октябр': 10, 'ноябр': 11, 'декабр': 12}
MONTH_RU = {1:'январь',2:'февраль',3:'март',4:'апрель',5:'май',6:'июнь',7:'июль',
            8:'август',9:'сентябрь',10:'октябрь',11:'ноябрь',12:'декабрь'}

def month_num(name):
    n = name.strip().lower()
    for k, v in MONTHS.items():
        if n.startswith(k): return v
    raise SystemExit(f'Не распознан месяц: «{name}»')

def parse_amount(s):
    # число с разрядами по 3 (иначе к сумме клеится цифра из «п.5.2.1» -> +10М). Разделитель — любой горизонт. пробел.
    m = re.search(r'(\d{1,3}(?:[^\S\n]\d{3})*),(\d{2})[^\S\n]*сум', s)
    if not m: return None
    return round(float(re.sub(r'[^\S\n]', '', m.group(1)) + '.' + m.group(2)), 2)
def read_docx_rows(path):
    """Возвращает список строк таблиц как (text). Только текст ячеек, по строкам."""
    try:
        root = ET.fromstring(zipfile.ZipFile(path).read('word/document.xml'))
    except Exception as e:
        raise SystemExit(f'Не удалось открыть docx «{path}»: {e}')
    body = root.find(f'{W}body')
    tx = lambda el: ''.join(t.text or '' for t in el.iter(f'{W}t'))
    rows = []
    for ch in body.iter():
        if ch.tag == f'{W}tr':
            rows.append(' | '.join(tx(tc).strip() for tc in ch.findall(f'{W}tc')))
    return rows

def read_pdf_rows(path):
    """PDF (новый формат отчёта Uzum) -> строки текста всех страниц."""
    import pdfplumber
    rows = []
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            rows.extend((pg.extract_text() or '').split('\n'))
    return rows

def parse_satori(path):
    """docx -> (month_int, year_int, {code: {'income':сум, 'startDay':int, 'days':int}})."""
    rows = read_pdf_rows(path) if path.lower().endswith('.pdf') else read_docx_rows(path)
    if not rows:
        raise SystemExit('В отчёте не найдено таблиц.')
    month = year = None
    acc = {}; order = []; cur = None
    fix_re = re.compile(r'(Fr\S+?-\d+)\s+за\s+с\s+[«"]?(\d{1,2})[»"]?\s+([А-Яа-яЁё]+)\s+(\d{4})', re.I)
    for ln in rows:
        mfix = fix_re.search(ln)
        amt = parse_amount(ln)
        if mfix:
            cur = mfix.group(1).upper()
            if month is None:
                month = month_num(mfix.group(3)); year = int(mfix.group(4))
            day = int(mfix.group(2))
            acc.setdefault(cur, {'income': 0.0, 'startDay': day})
            if amt: acc[cur]['income'] += amt
            order.append(cur)
        elif amt is not None and cur and re.search(r'п\.\s*1\.8|п\.\s*5\.2\.1|1\.8|5\.2\.1', ln):
            acc[cur]['income'] += amt
    if not acc:
        raise SystemExit('Не найдено строк «Фиксированное вознаграждение по Fr…». Неверный формат отчёта?')
    if month is None:
        raise SystemExit('Не удалось определить месяц/год отчёта.')
    dim = calendar.monthrange(year, month)[1]
    for c in acc:
        sd = acc[c]['startDay']
        acc[c]['days'] = dim - sd + 1
        acc[c]['income'] = round(acc[c]['income'], 2)
    # порядок как встречены
    ordered = {c: acc[c] for c in order}
    return month, year, dim, ordered

def read_ref_codes(path):
    wb = load_workbook(path, data_only=True); sh = wb['Справочник']
    cols = {str(sh.cell(4, j).value).strip(): j for j in range(1, sh.max_column + 1)}
    out = {}
    for rr in range(5, sh.max_row + 1):
        code = sh.cell(rr, cols['Код']).value
        if not code: continue
        out[str(code).strip()] = sh.cell(rr, cols['AKA']).value
    return out

def read_fact_sheet(wb):
    """Лист «Сатори_факт» -> (months[isoYYYY-MM], {month:{code:income}}, {month:{code:days}}). Пустой -> ({},{},{})."""
    if 'Сатори_факт' not in wb.sheetnames: return [], {}, {}
    sh = wb['Сатори_факт']
    keys = {}
    for j in range(3, sh.max_column + 1):
        v = sh.cell(3, j).value
        if v: keys[j] = str(v)
    if not keys: return [], {}, {}
    days = {}
    for j, m in keys.items():
        d = sh.cell(5, j).value; days[m] = int(d) if d is not None else 0
    inc = {m: {} for m in keys.values()}
    avg_r = None
    for rr in range(6, sh.max_row + 1):
        if str(sh.cell(rr, 1).value).strip() == 'ИТОГО по сети': avg_r = rr; break
    end = (avg_r - 1) if avg_r else sh.max_row
    for rr in range(6, end + 1):
        code = sh.cell(rr, 1).value
        if not code: continue
        for j, m in keys.items():
            v = sh.cell(rr, j).value
            if isinstance(v, (int, float)): inc[m][str(code).strip()] = v
    return sorted(keys.values()), inc, days

def write_fact_sheet(wb, months, inc, days_total, codes, ref):
    sheet = 'Сатори_факт'
    if sheet in wb.sheetnames:
        sh = wb[sheet]; sh.delete_rows(1, sh.max_row + 1)
    else:
        sh = wb.create_sheet(sheet)
    sh['A1'] = 'Факт по отчётам «Сатори» (агентское вознаграждение Uzum), сум — накопительно по месяцам'
    sh['A1'].font = Font(name='Arial', bold=True, size=12, color=NAVY)
    sh['A2'] = 'Доход ПВЗ = фикс. вознаграждение (гарантия, прорейчено) + п.1.8 + компенсация п.5.2.1. Закрытые месяцы не переписываются.'
    sh['A2'].font = Font(name='Arial', italic=True, size=9, color='595959')
    for i, m in enumerate(months): sh.cell(3, 3 + i, m)  # ISO ключ YYYY-MM
    sh.row_dimensions[3].hidden = True
    def mlabel(m):
        y, mo = m.split('-'); return f'{MONTH_RU[int(mo)]} {y}'
    hdr = ['Код', 'AKA'] + [mlabel(m) for m in months]
    for j, h in enumerate(hdr, 1):
        c = sh.cell(4, j, h); c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
        c.fill = PatternFill('solid', fgColor=NAVY); c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = BD
    c = sh.cell(5, 1, 'Дней в месяце'); c.font = Font(name='Arial', size=9, italic=True, color='595959'); c.border = BD
    sh.cell(5, 2, '').border = BD
    for i, m in enumerate(months):
        c = sh.cell(5, 3 + i, days_total.get(m, 0)); c.number_format = '0'
        c.font = Font(name='Arial', size=9, italic=True, color='595959'); c.alignment = Alignment(horizontal='center'); c.border = BD
    rr = 6
    for code in codes:
        sh.cell(rr, 1, code).border = BD; sh.cell(rr, 1).font = Font(name='Arial', size=10)
        sh.cell(rr, 2, ref.get(code, '')).border = BD; sh.cell(rr, 2).font = Font(name='Arial', size=10)
        for i, m in enumerate(months):
            v = inc[m].get(code)
            c = sh.cell(rr, 3 + i, round(v) if isinstance(v, (int, float)) else None)
            c.number_format = '#,##0'; c.font = Font(name='Arial', size=10); c.alignment = Alignment(horizontal='center'); c.border = BD
        rr += 1
    last = rr - 1
    c = sh.cell(rr, 1, 'ИТОГО по сети'); c.font = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=BLUE); c.border = BD
    sh.cell(rr, 2, '').fill = PatternFill('solid', fgColor=BLUE); sh.cell(rr, 2).border = BD
    for i in range(len(months)):
        L = get_column_letter(3 + i)
        c = sh.cell(rr, 3 + i, f'=SUM({L}6:{L}{last})'); c.number_format = '#,##0'
        c.font = Font(name='Arial', bold=True, color=WHITE, size=10); c.fill = PatternFill('solid', fgColor=BLUE); c.alignment = Alignment(horizontal='center'); c.border = BD
    for col, w in zip(['A', 'B'], [13, 22]): sh.column_dimensions[col].width = w
    for i in range(len(months)): sh.column_dimensions[get_column_letter(3 + i)].width = 15
    sh.freeze_panes = 'C6'

def main():
    month, year, dim, parsed = parse_satori(DOCX)
    iso = f'{year:04d}-{month:02d}'
    ref = read_ref_codes(REF)
    # сверка кодов со Справочником
    missing = [c for c in parsed if c not in ref]
    if missing:
        print('ВНИМАНИЕ: кодов нет в Справочнике (пропущены):', ', '.join(missing))
    wb = load_workbook(REF)
    months, inc, days = read_fact_sheet(wb)
    # накопление: новый месяц дописываем, существующий обновляем
    if iso not in months: months.append(iso)
    months = sorted(set(months))
    inc.setdefault(iso, {}); days[iso] = dim
    days_per_pvz = {}
    for code, d in parsed.items():
        if code not in ref: continue
        inc[iso][code] = d['income']; days_per_pvz[code] = d['days']
    codes = [c for c in ref]  # порядок Справочника
    write_fact_sheet(wb, months, inc, days, codes, ref)
    tmp_ref = REF + '.tmp'; wb.save(tmp_ref); os.replace(tmp_ref, REF)   # атомарная запись (без частичной порчи)
    total = sum(v for v in inc[iso].values())
    print(f'{REF}: лист «Сатори_факт» обновлён. Месяцев: {len(months)} ({months[0]} … {months[-1]}).')
    print(f'  Месяц {iso}: ПВЗ={len(inc[iso])}, ИТОГО доход={total:,.2f} сум')

    # ---- объект FACT в index.html: ВСЕ накопленные месяцы (доход по ПВЗ).
    #      Дни месяца и активные дни ПВЗ дашборд считает сам из даты старта Справочника. ----
    def _mlabel(m):
        yy, mm = m.split('-'); return f'{MONTH_RU[int(mm)]} {yy}'
    FACT = {'latest': months[-1], 'months': {m: {'label': _mlabel(m), 'income': inc[m]} for m in months}}
    # ---- FACT теперь в data.json (рядом с HTML); DATA/WK/SAL сохраняем как есть ----
    DJSON = os.path.join(os.path.dirname(os.path.abspath(HTML)) or '.', 'data.json')
    try: bundle = json.load(open(DJSON, encoding='utf-8'))
    except (OSError, ValueError): bundle = {}
    if not isinstance(bundle, dict): bundle = {}
    bundle['FACT'] = FACT
    bundle['updatedAt'] = dt.datetime.now().isoformat(timespec='seconds')   # метка свежести данных
    open(DJSON, 'w', encoding='utf-8').write(json.dumps(bundle, ensure_ascii=False))
    print(f'{DJSON}: FACT записан — месяцев {len(FACT["months"])} (последний {_mlabel(FACT["latest"])}).')
    print('Готово. Синхронизируй копию HTML, node --check, JSON.parse(data.json), затем git commit & push.')

if __name__ == '__main__':
    main()
